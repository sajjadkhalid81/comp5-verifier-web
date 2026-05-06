"""
verifier_core.py — Pure Python verification functions
Extracted from comp5_verifier.py — NO PyQt5 imports.
Used by both the web app (app.py) and can be used for testing.
"""

import os, re, io, zipfile, math, tempfile, subprocess
from pathlib import Path
from io import BytesIO

# ── Optional imports (graceful fallback) ─────────────────────────────────────
try:
    import fitz        # PyMuPDF
    PYMUPDF_OK = True
except ImportError:
    PYMUPDF_OK = False

try:
    from PIL import Image, ImageDraw, ImageFont, ImageEnhance
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from pypdf import PdfReader
    PYPDF_OK = True
except ImportError:
    PYPDF_OK = False

try:
    import numpy as np
    NUMPY_OK = True
except ImportError:
    NUMPY_OK = False

# ── Copy core functions from comp5_verifier.py ───────────────────────────────
# These are the pure-Python functions that do not depend on PyQt5

# REQUIRED_REVISION: default for backward compat only.
# Actual revision check always uses the Excel row value.
REQUIRED_REVISION = "B"  # legacy — not used in check_revision()
CLASSIF_VARIANTS  = [
    "Classification: Internal", "Classification:Internal",
    "CLASSIFICATION: INTERNAL", "classification: internal",
]

COORDS = {
    "A1": {
        "doc_no":   (2110, 1463, 2350, 1483),
        "cpy_no":   (2053, 1598, 2225, 1618),
        "revision": (2290, 1598, 2340, 1618),
        "sigs": {
            "DRN":   (2195, 1130, 2232, 1185),
            "DSGND": (2232, 1130, 2264, 1185),
            "CHKD":  (2264, 1130, 2295, 1185),
            "APVD1": (2295, 1130, 2325, 1185),
            "APVD2": (2325, 1130, 2362, 1185),
        },
    }
}

CLASSIF_REGION = (0, 0, 500, 80)
DOC_NO_PATTERN = re.compile(r"\d{4,6}-[A-Z]-[A-Z0-9]+-\d+-[A-Z]+-[A-Z]+-[A-Z]+-\d+-\d+")
CPY_RE         = re.compile(r"(\d{3}-\d{2}-[A-Z]+-[A-Z]+-\d{4,5})", re.IGNORECASE)


def normalize(s):
    return re.sub(r"[\s\-_]", "", str(s)).upper()


def normalize_cpy_seq(s):
    s = str(s).strip()
    m = re.match(r"^(\d{3}-\d{2}-[A-Za-z]+-[A-Za-z]+-)(\d{4,5})$", s)
    if m:
        return m.group(1) + m.group(2).zfill(5)
    return s


def doc_no_match(pdf_val, excel_val):
    if not pdf_val or not excel_val:
        return False
    p = normalize(pdf_val)
    e = normalize(excel_val)
    if p == e:
        return True
    def split_prefix(s):
        m = re.match(r"^(\d+)([A-Z].+)$", s)
        return (m.group(1), m.group(2)) if m else (s, "")
    p_num, p_rest = split_prefix(p)
    e_num, e_rest = split_prefix(e)
    if p_rest and e_rest and p_rest == e_rest:
        if e_num.endswith(p_num) or p_num.endswith(e_num):
            return True
    return False


def cpy_no_match(pdf_val, excel_val):
    if not pdf_val or not excel_val:
        return False
    n_pdf   = re.sub(r"[\s\-_]", "", pdf_val).upper()
    n_excel = re.sub(r"[\s\-_]", "", excel_val).upper()
    if n_pdf == n_excel:
        return True
    return (re.sub(r"[\s\-_]", "", normalize_cpy_seq(pdf_val)).upper() ==
            re.sub(r"[\s\-_]", "", normalize_cpy_seq(excel_val)).upper())


def collect_pdfs_from_zip(zip_source, depth=0):
    """Recursively collect PDFs from ZIP (any nesting depth)."""
    results = []
    if depth > 5:
        return results
    try:
        if isinstance(zip_source, (str, Path)):
            zf_obj = zipfile.ZipFile(zip_source, "r")
        else:
            zf_obj = zipfile.ZipFile(BytesIO(zip_source), "r")
        with zf_obj as zf:
            for name in sorted(n for n in zf.namelist()
                               if not n.startswith("__") and not n.endswith("/")):
                short = name.split("/")[-1]
                low   = name.lower()
                if low.endswith(".pdf"):
                    results.append((short, zf.read(name)))
                elif low.endswith(".zip"):
                    results.extend(collect_pdfs_from_zip(zf.read(name), depth + 1))
    except Exception:
        pass
    seen = set()
    unique = []
    for fname, data in results:
        if fname not in seen:
            seen.add(fname)
            unique.append((fname, data))
    return unique


def load_transmittal_excel(excel_source):
    """
    Load transmittal from Excel file (path or bytes).
    Returns list of dicts with keys: srNo, docNo, cpyNo, revision, title.
    Reads by Document No. — not by Sr. No. — so blank rows never stop reading.
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl not installed")

    if isinstance(excel_source, (str, Path)):
        wb = openpyxl.load_workbook(excel_source, data_only=True)
    else:
        wb = openpyxl.load_workbook(BytesIO(excel_source), data_only=True)

    sheet_name = next(
        (n for n in wb.sheetnames if "QG-NFPS" in n.upper()),
        wb.sheetnames[0]
    )
    ws = wb[sheet_name]

    transmittal = []
    sr_counter  = 0
    for row in ws.iter_rows(min_row=25, values_only=True):
        doc_no = str(row[1] or "").strip()
        cpy_no = str(row[2] or "").strip()
        if not doc_no and not cpy_no:
            continue
        if doc_no.lower() in ("document no.", "doc no") or \
           cpy_no.lower() in ("client doc. no.", "cpy no"):
            continue

        # Skip rows where Doc No. looks like a date (Excel date cell misread)
        # e.g. "2026-04-19 00:00:00" — happens when Excel has a date in that column
        _date_pat = re.compile(r"^\d{4}-\d{2}-\d{2}(\s|T|$)")
        if _date_pat.match(doc_no) or _date_pat.match(cpy_no):
            continue

        # Skip rows with no valid drawing number pattern in either column
        _cpy_valid = re.compile(r"\d{3}-\d{2}-[A-Za-z]+-[A-Za-z]+-\d{4,5}")
        _doc_valid = re.compile(r"\d{4,6}-[A-Za-z]-[A-Za-z0-9]+-\d+-[A-Za-z]+-[A-Za-z]+-[A-Za-z]+-\d+-\d+")
        if not _cpy_valid.search(cpy_no) and not _doc_valid.search(doc_no):
            continue
        sr_counter += 1
        try:
            sr = int(float(str(row[0]).strip())) if row[0] else sr_counter
        except (ValueError, TypeError):
            sr = sr_counter
        transmittal.append({
            "srNo":     sr,
            "docNo":    doc_no,
            "cpyNo":    cpy_no,
            "revision": str(row[3] or "").strip().replace("\xa0", "").strip(),
            "title":    str(row[4] or "").strip(),
        })
    return transmittal


def get_page_type(page):
    w, h = page.rect.width, page.rect.height
    if abs(w - 2384) < 200 and abs(h - 1684) < 200:
        return "A1"
    if abs(w - 2551) < 200 and abs(h - 1772) < 200:
        return "A1"  # HVAC variant
    if abs(w - 792) < 50 and abs(h - 612) < 50:
        return "AB"
    if abs(w - 595) < 50 and abs(h - 842) < 50:
        return "A4"
    return "A1"  # default


def extract_text_at(page, coords):
    x0, top, x1, bot = coords
    return page.get_text("text", clip=fitz.Rect(x0, top, x1, bot)).strip()


def verify_pdf(pdf_bytes, filename, row):
    """
    Full 8-check verification of a single PDF.
    Returns result dict compatible with web and desktop versions.
    """
    if not PYMUPDF_OK:
        return _error_result(filename, row, "PyMuPDF not installed")

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        return _error_result(filename, row, f"Cannot open PDF: {e}")

    try:
        page      = doc[0]
        page_type = get_page_type(page)
        excel_doc = str(row.get("docNo", "")).strip()
        excel_cpy = str(row.get("cpyNo", "")).strip()
        excel_rev = str(row.get("revision", "")).strip()
        excel_ttl = str(row.get("title", "")).strip()

        # 1 ── Doc No
        doc_status, doc_display, doc_from_pdf = _check_doc_no(page, page_type, excel_doc)

        # 2 ── CPY No
        cpy_status, cpy_display, cpy_from_pdf = _check_cpy_no(page, page_type, filename, excel_cpy)

        # 3 ── Revision
        rev_status, rev_display, rev_from_pdf = _check_revision(page, page_type, filename, excel_rev)

        # 4 ── Signatures
        sig_status, sig_count = _check_signatures(doc, page_type)

        # 5 ── Comments
        com_status, com_count = _check_comments(pdf_bytes)

        # 6 ── Classification
        cls_status, cls_missing = _check_classification(doc)

        # 7 ── Prev Rev
        prev_status = _check_prev_rev(page)

        # 8 ── Title
        ttl_status, ttl_display = _check_title(page, excel_ttl)

        # 9 ── Multi-sheet check (sheets 2+ if PDF has multiple pages)
        # Each sheet has its own title block — verify Doc No, CPY No, Revision
        # on every sheet, not just sheet 1.
        # Signatures/Prev Rev: only on sheet 1 — no change needed.
        # Classification: already checked across all pages by _check_classification(doc).
        sheet_issues = []
        if len(doc) > 1:
            for sheet_idx in range(1, len(doc)):
                pg = doc[sheet_idx]
                pg_type = get_page_type(pg)
                sheet_num = sheet_idx + 1

                # Doc No on this sheet
                _, _, pg_doc = _check_doc_no(pg, pg_type, excel_doc)
                if pg_doc and excel_doc and not doc_no_match(pg_doc, excel_doc):
                    sheet_issues.append(
                        f"Sheet {sheet_num}: Doc No mismatch (PDF: {pg_doc} | Excel: {excel_doc})"
                    )

                # CPY No on this sheet
                _, _, pg_cpy = _check_cpy_no(pg, pg_type, filename, excel_cpy)
                if pg_cpy and excel_cpy and not cpy_no_match(pg_cpy, excel_cpy):
                    sheet_issues.append(
                        f"Sheet {sheet_num}: CPY No mismatch (PDF: {pg_cpy} | Excel: {excel_cpy})"
                    )

                # Revision on this sheet — use coord extraction only (filename is same for all sheets)
                raw_rev = extract_text_at(pg, COORDS["A1"]["revision"]).strip()
                if raw_rev:
                    pg_rev = raw_rev.upper()
                    if excel_rev and pg_rev != excel_rev.upper():
                        sheet_issues.append(
                            f"Sheet {sheet_num}: Rev mismatch (PDF: {pg_rev} | Excel: {excel_rev})"
                        )

        doc.close()

        hard_fail = any(s == "FAIL" for s in [
            rev_status, sig_status, com_status, cls_status, doc_status, cpy_status
        ])
        # Sheet mismatches across pages = FAIL
        if sheet_issues:
            hard_fail = True

        all_pass = all(s == "PASS" for s in [
            doc_status, cpy_status, rev_status, sig_status, com_status,
            cls_status, prev_status, ttl_status
        ]) and not sheet_issues

        overall = "FAIL" if hard_fail else ("PASS" if all_pass else "WARN")

        issues_parts = []
        if doc_status == "FAIL": issues_parts.append(f"Doc No mismatch (PDF: {doc_from_pdf} | Excel: {excel_doc})")
        if cpy_status == "FAIL": issues_parts.append(f"CPY No mismatch (PDF: {cpy_from_pdf} | Excel: {excel_cpy})")
        if rev_status == "FAIL": issues_parts.append(f"Rev mismatch (PDF: {rev_from_pdf})")
        if sig_status == "FAIL": issues_parts.append(f"Insufficient signatures — found {sig_count}, required 3 (DRN/BY + CHKD + APVD)")
        if com_status == "FAIL": issues_parts.append(f"{com_count} comment(s)")
        if cls_status == "FAIL": issues_parts.append(f"Classification missing on pages: {cls_missing}")
        if cls_status == "WARN": issues_parts.append(f"Classification not found on pages: {cls_missing}")
        if prev_status == "WARN": issues_parts.append("Prev rev not confirmed")
        if ttl_status == "WARN": issues_parts.append("Title not confirmed")
        issues_parts.extend(sheet_issues)

        return {
            "filename":    filename,
            "srNo":        row.get("srNo", ""),
            "docNo":       excel_doc,
            "cpyNo":       excel_cpy,
            "revision":    excel_rev,
            "title":       excel_ttl,
            "docNoFromPdf":  doc_from_pdf,
            "cpyNoFromPdf":  cpy_from_pdf,
            "revFromPdf":    rev_from_pdf,
            "titleFromPdf":  ttl_display,
            "docNoMatch":    doc_status,
            "cpyNoMatch":    cpy_status,
            "revMatch":      rev_status,
            "sigsResult":    sig_status,
            "sigCount":      sig_count,
            "commentsResult":      com_status,
            "commentsCount":       com_count,
            "classificationResult":      cls_status,
            "classificationMissingPages": cls_missing,
            "prevRevResult": prev_status,
            "titleMatch":    ttl_status,
            "overallResult": overall,
            "issues":        "; ".join(issues_parts) if issues_parts else "None",
        }

    except Exception as e:
        try:
            doc.close()
        except Exception:
            pass
        return _error_result(filename, row, str(e))


def _error_result(filename, row, error_msg):
    return {
        "filename": filename, "srNo": row.get("srNo",""),
        "docNo": row.get("docNo",""), "cpyNo": row.get("cpyNo",""),
        "revision": row.get("revision",""), "title": row.get("title",""),
        "docNoFromPdf": "", "cpyNoFromPdf": "", "revFromPdf": "", "titleFromPdf": "",
        "docNoMatch": "FAIL", "cpyNoMatch": "FAIL", "revMatch": "FAIL",
        "sigsResult": "FAIL", "sigCount": 0,
        "commentsResult": "FAIL", "commentsCount": 0,
        "classificationResult": "FAIL", "classificationMissingPages": [],
        "prevRevResult": "FAIL", "titleMatch": "FAIL",
        "overallResult": "FAIL",
        "issues": f"Error: {error_msg}",
    }


def _check_doc_no(page, page_type, excel_val):
    # Try coordinate extraction first (works for standard A1)
    raw = extract_text_at(page, COORDS["A1"]["doc_no"])
    m = DOC_NO_PATTERN.search(raw) if raw else None
    pdf_val = m.group(0) if m else ""

    # Fallback: search full page text (works for rotated/non-A1 drawings)
    if not pdf_val:
        full = page.get_text("text").replace("\x1f", " ")
        # Look for "CONTRACTOR DRAWING NO:" label then the number after it
        m2 = re.search(
            r"CONTRACTOR\s+DRAWING\s+NO[:\s]+(" + DOC_NO_PATTERN.pattern + r")",
            full, re.IGNORECASE
        )
        if not m2:
            m2 = DOC_NO_PATTERN.search(full)
        pdf_val = m2.group(1) if (m2 and m2.lastindex) else (m2.group(0) if m2 else "")

    if pdf_val:
        if not excel_val:
            return "PASS", f"✓ {pdf_val}", pdf_val
        if doc_no_match(pdf_val, excel_val):
            return "PASS", f"✓ {pdf_val}", pdf_val
        return "FAIL", pdf_val, pdf_val
    return "PASS", f"✓ {excel_val}  |  Remark: vector font — verify manually", excel_val


def _find_by_label(page, label_variants, value_pattern):
    """
    Universal label-proximity extraction.
    Works for any drawing type regardless of coordinate system or rotation.
    Detects text direction (normal/rotated) from character-level data.

    Also searches INSIDE the label block itself — handles drawings where
    label and value are in the same text block e.g.:
    "CONTRACTOR DRAWING NO: 033784-E-FAX-24-SPM-AR-D-10027-0001"
    """
    words    = page.get_text("words")
    rawdict  = page.get_text("rawdict")

    # Build origin→direction lookup from character-level data
    char_dirs = {}
    for blk in rawdict.get("blocks", []):
        if blk.get("type") != 0: continue
        for line in blk.get("lines", []):
            d = line.get("dir", (1, 0))
            for span in line.get("spans", []):
                for ch in span.get("chars", []):
                    ox, oy = ch["origin"]
                    char_dirs[(round(ox), round(oy))] = d

    def word_dir(wx0, wy0):
        return char_dirs.get((round(wx0), round(wy0)), (1, 0))

    for variant in label_variants:
        var_words = variant.upper().replace(":", "").replace(".", "").split()
        for i, w in enumerate(words):
            if w[4].upper().replace(":", "").replace(".", "") != var_words[0]:
                continue
            # Match all label words
            ok = True
            for j, vw in enumerate(var_words[1:], 1):
                if i+j >= len(words) or words[i+j][4].upper().replace(":","").replace(".","") != vw:
                    ok = False; break
            if not ok: continue

            span = words[i:i+len(var_words)]
            lx0 = min(s[0] for s in span); ly0 = min(s[1] for s in span)
            lx1 = max(s[2] for s in span); ly1 = max(s[3] for s in span)

            d = word_dir(w[0], w[1])
            is_rotated = abs(d[1]) > abs(d[0])

            candidates = []
            for w2 in words:
                wx0,wy0,wx1,wy1,wtext = w2[:5]
                if not wtext.strip() or w2 in span: continue
                wcx=(wx0+wx1)/2; wcy=(wy0+wy1)/2

                if is_rotated:
                    # Rotated text: value column is at higher x, overlapping y
                    y_overlap = not (wy1 < ly0-10 or wy0 > ly1+10)
                    if wx0 > lx1-5 and y_overlap and (wx0-lx1) < 200:
                        candidates.append((wx0-lx1, wtext))
                else:
                    # Normal text: value is to right or directly below
                    if abs(wcy-(ly0+ly1)/2) < 20 and wx0 >= lx1-5:
                        candidates.append((wx0-lx1, wtext))
                    elif wy0 >= ly1-5 and abs(wcx-(lx0+lx1)/2) < 120:
                        candidates.append((wy0-ly1+500, wtext))

            if candidates:
                candidates.sort(key=lambda x: x[0])
                window = " ".join(c[1] for c in candidates[:20])
                m = value_pattern.search(window)
                if m:
                    return m.group(1)
    # Fallback: search inside each label block (label + value in same block)
    # e.g. "CONTRACTOR DRAWING NO: 033784-..."
    blocks = page.get_text("blocks")
    for variant in label_variants:
        var_upper = variant.upper()
        for b in blocks:
            if var_upper in b[4].upper().replace("\x1f", " ").replace(":", "").replace(".", ""):
                m = value_pattern.search(b[4].replace("\x1f", " ").replace("\n", " "))
                if m:
                    return m.group(1)

    return None


def _check_cpy_no(page, page_type, filename, excel_val):
    """
    Intelligent CPY extraction — coordinate-free.
    Uses label-proximity search then exact-match fallback.
    Works for all drawing types: architectural, P&ID, structural, rotated.
    """
    fname_cpy = re.sub(r"_[A-Z]\.pdf$", "", filename, flags=re.IGNORECASE)
    fname_cpy = re.sub(r"\.pdf$", "", fname_cpy, flags=re.IGNORECASE)

    LABELS = ["PROJECT DRAWING NO", "DRAWING NO", "CLIENT DOCUMENT NO",
              "CLIENT DOC NO", "PROJ DWG NO"]

    # Step 1: label-proximity search
    pdf_val = _find_by_label(page, LABELS, CPY_RE)

    # Step 2: if label found wrong value (mismatch with excel), try exact match
    # This handles P&ID drawings where "DRAWING NO." appears in valve tables too
    if excel_val and pdf_val:
        if not cpy_no_match(pdf_val, excel_val):
            pdf_val = None  # discard wrong label result → fall to exact match

    # Step 3: exact match — find the CPY pattern that matches excel_val exactly
    if not pdf_val and excel_val:
        full = page.get_text("text").replace("\x1f", " ")
        norm_expected = re.sub(r"[-]", "", str(excel_val)).upper()
        for c in CPY_RE.findall(full):
            if re.sub(r"[-]", "", c).upper() == norm_expected:
                pdf_val = c
                break

    # Step 4: filename fallback
    if not pdf_val:
        pdf_val = fname_cpy or None

    # Compare and return
    if pdf_val:
        if not excel_val:
            return "WARN", pdf_val, pdf_val
        if cpy_no_match(pdf_val, excel_val):
            return "PASS", f"\u2713 {pdf_val}", pdf_val
        return "FAIL", pdf_val, pdf_val

    return "PASS", f"\u2713 {excel_val}  |  Remark: verify manually", excel_val



def _check_revision(page, page_type, filename, excel_val):
    """
    Compare PDF revision against the Excel transmittal value.
    Uses Excel value — NOT a hardcoded constant — so any revision letter works.
    P&ID drawings at Rev A, architectural at Rev B, etc. all handled correctly.
    """
    expected = str(excel_val or "").strip().upper()
    if not expected:
        expected = REQUIRED_REVISION  # fallback if Excel cell is blank

    # Method 1: revision letter from filename (most reliable — e.g. _A.pdf, _B.pdf)
    m = re.search(r"_([A-Za-z])\.pdf$", filename, re.IGNORECASE)
    if m:
        rev = m.group(1).upper()
        if rev == expected:
            return "PASS", f"✓ {rev}", rev
        return "FAIL", rev, rev

    # Method 2: coordinate extraction from title block (A1 standard drawings)
    raw = extract_text_at(page, COORDS["A1"]["revision"]).strip()
    if raw:
        rev = raw.strip().upper()
        if rev == expected:
            return "PASS", f"✓ {rev}", rev
        return "FAIL", rev, rev

    # Method 3: full-page text search for revision letter near SHT/REVISION block
    full = page.get_text("text").replace("\x1f", " ")
    idx = full.upper().find("SHT")
    if idx >= 0:
        ctx = full[max(0, idx-20):idx+80]
        # Pattern: "200-20-PR-PID-00182 0001 A" — last token before next block
        tokens = ctx.split()
        for tok in reversed(tokens):
            if re.match(r"^[A-Z]$", tok.upper()):
                rev = tok.upper()
                if rev == expected:
                    return "PASS", f"✓ {rev}", rev
                return "FAIL", rev, rev

    return "WARN", f"?/{expected}", ""


def _detect_rotation(page):
    """
    Majority-vote rotation detection.
    Returns True only if MOST characters are rotated (like P&ID drawings).
    Prevents false rotated detection from classification stamps or watermarks.
    """
    rawdict = page.get_text("rawdict")
    normal = rotated = 0
    for blk in rawdict.get("blocks", []):
        if blk.get("type") != 0:
            continue
        for line in blk.get("lines", []):
            d = line.get("dir", (1, 0))
            chars = sum(len(s.get("chars", [])) for s in line.get("spans", []))
            if abs(d[0]) >= abs(d[1]):
                normal += chars
            else:
                rotated += chars
    return rotated > normal


def _has_sig_content(text):
    """
    Returns True if text represents an actual signature.
    Returns False for empty cells, "NOT REQUIRED", "N/A" etc.
    """
    clean = " ".join(text.strip().split())
    if not clean or len(clean) < 2:
        return False
    # Remove "NOT REQUIRED" fragments before checking
    clean2 = re.sub(r"not\s+required|not\s+req", "", clean, flags=re.IGNORECASE).strip()
    if len(clean2) < 2:
        return False
    # Initials (2-5 uppercase letters)
    if re.search(r"\b[A-Z]{2,5}\b", clean2):
        return True
    # Date patterns
    if re.search(r"\d{2}[-./]\d{2}|\d{4}", clean2):
        return True
    # Person names (mixed case)
    if re.search(r"[A-Z][a-z]{2,}", clean2):
        return True
    return False


def _check_signatures(doc, page_type):
    """
    Universal signature check — label-driven, no fixed coordinates.

    Handles all drawing types and signature styles:
    - AB Architectural: DRN / CHKD / APVD labels in horizontal row
    - PR P&ID / PI:     BY  / CHKD / APVD labels in vertical (rotated) column
    - HV / ST:          Scanned sigs — person names + dates clustered together

    Method 1 — Label proximity (DRN/BY + CHKD + APVD):
        Finds all sig labels, groups by row (normal) or column (rotated).
        Reads text ABOVE (normal) or RIGHT (rotated) of each label.
        Counts cells with actual content vs "NOT REQUIRED" / empty.

    Method 2 — Digital sig text (Foxit visual signatures):
        Counts DN: CN= entries and "Digitally signed by" phrases.

    Method 3 — Date cluster (scanned signatures):
        Finds date blocks (20XX.MM pattern) clustered in same area.
        Count of clustered dates = count of signers.

    Required: >= 3 signed cells for PASS.
    """
    page = doc[0]
    words  = page.get_text("words")
    full   = page.get_text("text").replace("\x1f", " ")
    blocks = page.get_text("blocks")

    SIG_LABELS = {"BY", "DRN", "CHKD", "APVD", "DSGND"}
    REQUIRED   = {"BY", "DRN", "CHKD", "APVD"}

    is_rotated = _detect_rotation(page)

    # ── Method 1: Label-proximity ─────────────────────────────────────────────
    from collections import defaultdict
    rows = defaultdict(list)
    for w in words:
        clean = w[4].upper().replace(":", "").replace(".", "").strip()
        if clean in SIG_LABELS:
            # Group by y (normal) or x (rotated)
            key = round((w[0] if is_rotated else w[1]) / 8) * 8
            rows[key].append((clean, w[0], w[1], w[2], w[3]))

    label_signed = 0
    for key, labels in rows.items():
        label_names = {l[0] for l in labels}
        # Only process rows that have both CHKD and APVD (required pair)
        if "CHKD" not in label_names or "APVD" not in label_names:
            continue
        signed_count = 0
        for lname, lx0, ly0, lx1, ly1 in labels:
            if lname not in REQUIRED:
                continue
            if is_rotated:
                # Content to the RIGHT of label in rotated drawings
                content = page.get_text("text",
                    clip=fitz.Rect(lx1, ly0-30, lx1+120, ly1+30)
                ).strip().replace("\n", " ")
            else:
                # Content ABOVE label in normal drawings
                content = page.get_text("text",
                    clip=fitz.Rect(lx0-3, max(0, ly0-80), lx1+3, ly0)
                ).strip().replace("\n", " ")
            if _has_sig_content(content):
                signed_count += 1
        label_signed = max(label_signed, signed_count)

    # ── Method 2: Digital sig text (Foxit DN: CN=) ───────────────────────────
    dns        = re.findall(r"DN:\s*CN=([^,\n]+)", full)
    signed_by  = re.findall(r"[Dd]igitally\s+signed\s+by\s+(\S+)", full)
    digital_signed = max(len(dns), len(signed_by))

    # ── Method 3: Date cluster (scanned sigs) ────────────────────────────────
    # Relaxed pattern handles dates split across lines: "2026.04.\n13"
    DATE_PAT = re.compile(r"20\d{2}[.\-/]\d{2}")
    date_blocks = []
    for b in blocks:
        if b[6] != 0:
            continue
        txt = b[4].replace("\x1f", " ").replace("\x00", " ").replace("\n", " ")
        if DATE_PAT.search(txt):
            date_blocks.append(b[1])  # store y-position

    cluster_signed = 0
    if len(date_blocks) >= 2:
        y_range = max(date_blocks) - min(date_blocks)
        if y_range < 150:  # all within 150pts = same sig block
            cluster_signed = len(date_blocks)

    # ── Take highest count from all methods ──────────────────────────────────
    total  = max(label_signed, digital_signed, cluster_signed)
    status = "PASS" if total >= 3 else "FAIL"
    # Cap display at 3 — required is 3, showing more is confusing
    display_count = min(total, 3)
    return status, display_count


def _find_title_block_region(page):
    """Dynamically locate title block bounding box using anchor labels."""
    words = page.get_text("words")
    pw, ph = page.rect.width, page.rect.height
    ANCHORS = [
        ["CONTRACTOR", "DRAWING", "NO"], ["PROJECT", "NO"],
        ["REVISION"], ["SCALE"], ["SHT"], ["DRN"], ["CHKD"], ["APVD"],
    ]
    positions = []
    for anchor in ANCHORS:
        for i, w in enumerate(words):
            if w[4].upper().replace(":", "").replace(".", "") != anchor[0]:
                continue
            ok = all(
                i+j < len(words) and
                words[i+j][4].upper().replace(":", "").replace(".", "") == av
                for j, av in enumerate(anchor[1:], 1)
            )
            if ok:
                span = words[i:i+len(anchor)]
                positions.append((
                    (min(s[0] for s in span) + max(s[2] for s in span)) / 2,
                    (min(s[1] for s in span) + max(s[3] for s in span)) / 2,
                ))
                break
    if len(positions) < 2:
        return None
    xs = [p[0] for p in positions]
    ys = [p[1] for p in positions]
    margin = 80
    return (max(0, min(xs)-margin), max(0, min(ys)-margin),
            min(pw, max(xs)+margin), min(ph, max(ys)+margin))


def _check_comments(pdf_bytes):
    """Check for reviewer annotations/comments in the PDF."""
    if not PYPDF_OK:
        return "WARN", 0
    try:
        reader = PdfReader(BytesIO(pdf_bytes))
        count  = 0
        for pg in reader.pages:
            annots = pg.get("/Annots")
            if annots:
                for ann in annots:
                    try:
                        ao  = ann.get_object()
                        sub = str(ao.get("/Subtype", "")).replace(" ", "")
                        if sub not in ("/Widget", "/Link"):
                            count += 1
                    except Exception:
                        pass
        return ("PASS" if count == 0 else "FAIL"), count
    except Exception:
        return "WARN", 0


def _check_classification(doc):
    """
    Fully intelligent classification check — no pixels, no coordinates.

    Uses the same label-proximity mechanism as Doc No, CPY No, Revision.

    Algorithm:
      1. Fast full-text search for explicit "Classification: Internal" phrase
      2. Dynamically locate title block using anchor labels
         (CONTRACTOR DRAWING NO, REVISION, SCALE, etc.)
      3. Search for "Classification" label ONLY within title block region
         — eliminates false positives from drawing body content
         (e.g. "ARCHITECTURAL INTERNAL DOORS DETAIL" in notes)
      4. Read adjacent value using label-proximity logic

    Results:
      PASS — "Classification: Internal" confirmed in title block
      WARN — No classification field in title block (template has none)
              OR label found but value is not "Internal"
    """
    missing = []

    for i, page in enumerate(doc):
        words   = page.get_text("words")
        full    = page.get_text("text").replace("\x1f", " ")
        compact = re.sub(r"\s+", " ", full)

        # Step 1: Fast full-text phrase match
        if re.search(r"classif.{0,5}ication\s*:\s*internal", compact, re.IGNORECASE):
            continue  # PASS

        # Step 2: Locate title block dynamically
        tb = _find_title_block_region(page)

        # Step 3: Search for "Classification" label INSIDE title block only
        cls_word = None
        for w in words:
            if w[4].replace(":", "").strip().upper() != "CLASSIFICATION":
                continue
            if tb:
                # Reject if this word is outside the title block region
                wcx = (w[0] + w[2]) / 2
                wcy = (w[1] + w[3]) / 2
                if not (tb[0] <= wcx <= tb[2] and tb[1] <= wcy <= tb[3]):
                    continue
            cls_word = w
            break

        if not cls_word:
            missing.append(i + 1)
            continue

        # Step 4: Read value adjacent to label
        lx0, ly0, lx1, ly1 = cls_word[:4]
        candidates = []
        for w2 in words:
            wx0, wy0, wx1, wy1, wtext = w2[:5]
            if not wtext.strip() or w2 == cls_word:
                continue
            wcx, wcy = (wx0+wx1)/2, (wy0+wy1)/2
            # Right of label (normal orientation)
            if abs(wcy-(ly0+ly1)/2) < 25 and wx0 >= lx1-5:
                candidates.append((wx0-lx1, wtext))
            # Below label
            elif wy0 >= ly1-5 and (wy0-ly1) < 40 and abs(wcx-(lx0+lx1)/2) < 80:
                candidates.append((wy0-ly1+500, wtext))
            # Rotated (higher x, overlapping y — P&ID rotated title block)
            elif wx0 > lx1-5 and not (wy1 < ly0-10 or wy0 > ly1+10) and (wx0-lx1) < 150:
                candidates.append((wx0-lx1+1000, wtext))

        if candidates:
            candidates.sort(key=lambda x: x[0])
            value = " ".join(c[1] for c in candidates[:5]).strip()
            if "INTERNAL" in value.upper():
                continue  # PASS
        # Label found but value unreadable or not "Internal"
        missing.append(i + 1)

    return ("PASS" if not missing else "WARN"), missing


def _check_prev_rev(page):
    # Replace  (PDF word separator) with space before searching
    # Without this, "ISSUEDFORINTERDISCIPLINE" does not match "ISSUED FOR INTER"
    text = page.get_text("text").replace("", " ").upper()
    phrases = [
        "INTER-DISCIPLINE CHECK", "INTER DISCIPLINE CHECK",
        "ISSUED FOR INTER", "INTERDISCIPLINE CHECK",
        "INTERDISCIPLINE",
    ]
    if any(p in text for p in phrases):
        return "PASS"
    return "WARN"


def _check_title(page, expected):
    if not expected:
        return "WARN", "(no title in Excel)"
    text  = page.get_text("text")
    words = {w for w in re.sub(r"[^A-Za-z0-9 ]", " ", expected.upper()).split() if len(w) > 2}
    found = {w for w in re.sub(r"[^A-Za-z0-9 ]", " ", text.upper()).split()       if len(w) > 2}
    if not words:
        return "WARN", "(empty title)"
    pct = len(words & found) / len(words)
    if pct >= 0.70:
        return "PASS", expected
    if pct >= 0.50:
        return "PASS", f"{expected}  |  Remark: partial {int(pct*100)}% match"
    return "WARN", expected


def generate_excel_report(results, transmittal_name=""):
    """Generate Excel report bytes from results list."""
    if not OPENPYXL_OK:
        raise ImportError("openpyxl not installed")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verification"
    headers = ["Sr.", "Filename", "Doc No (Excel)", "Doc No Match",
               "CPY Match", "Rev Match", "Signatures", "Comments",
               "Classification", "Prev Rev", "Title", "RESULT", "Issues"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(bold=True, color="FFFFFF")
    for ri, r in enumerate(results, 2):
        try:
            ov = r.get("overallResult", "WARN")
            sig_count = r.get("sigCount", 0)
            sig_res   = r.get("sigsResult", "WARN")
            ws.cell(ri, 1,  r.get("srNo", ri-1))
            ws.cell(ri, 2,  r.get("filename", ""))
            ws.cell(ri, 3,  r.get("docNo", ""))
            ws.cell(ri, 4,  r.get("docNoMatch", "WARN"))
            ws.cell(ri, 5,  r.get("cpyNoMatch", "WARN"))
            ws.cell(ri, 6,  r.get("revMatch", "WARN"))
            ws.cell(ri, 7,  f"{sig_count}/3  {sig_res}")
            ws.cell(ri, 8,  r.get("commentsResult", "WARN"))
            ws.cell(ri, 9,  r.get("classificationResult", "WARN"))
            ws.cell(ri, 10, r.get("prevRevResult", "WARN"))
            ws.cell(ri, 11, r.get("titleMatch", "WARN"))
            ws.cell(ri, 12, ov)
            ws.cell(ri, 13, str(r.get("issues", "")))
            color = "1E8449" if ov=="PASS" else "C0392B" if ov=="FAIL" else "D68910"
            ws.cell(ri, 12).fill = PatternFill("solid", fgColor=color)
            ws.cell(ri, 12).font = Font(bold=True, color="FFFFFF")
        except Exception:
            pass   # never let one bad row break the whole report
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
